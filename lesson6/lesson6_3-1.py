from pprint import pprint

import openpyxl.workbook
import openpyxl.worksheet
import tools
import openpyxl
from openpyxl import load_workbook, Workbook, worksheet


def main():
    sitenames = tools.get_sitenames('aqi.xlsx')
    print(sitenames)
    wb:Workbook = openpyxl.Workbook()
    sheet:worksheet = wb.active
    sheet.title = "站點名稱"
    for idy , name in enumerate(sitenames):
        print(name)
        sheet.cell(column=1,row=idy+1, value=name)
    wb.save('Sitename.xlsx')



if __name__ == '__main__':
    main()